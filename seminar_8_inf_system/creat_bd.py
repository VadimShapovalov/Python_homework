"""
Модуль генерации первоночальной бд с данными
"""

from logger import operation_logger as logg
import openpyxl

def cr_db():
    wb = openpyxl.Workbook()
    wb.create_sheet(title='Сотрудники', index=0)
    wb.active[f"A1"].value = "Номер сотрудника"
    wb.active[f"B1"].value = "Имя"
    wb.active[f"C1"].value = "Фамилия"
    wb.active[f"D1"].value = "Телефон"
    wb.active[f"E1"].value = "Описание"
    wb.active[f"F1"].value = "Статус"
    ls = [[1, "Даша", "Тараканова", "+79324733228", "-", 1],
        [2, "Марина", "Ковалева", "+79324733229", "-", 1],
        [3, "Илья", "Рыбаков", "+ 79324733228", "-", 1],
        [4, "Валентина", "Сорокина", "+ 79324733228", "-", 1],
        [5, "Денис", "Орлов", "+ 79324733228", "-", 1],
        [6, "Андрей", "Славнов", "+ 79324733228", "-", 0],]
    for i in range(2, 8):
        wb.active[f"A{i}"].value = ls[i-2][0]
        wb.active[f"B{i}"].value = ls[i-2][1]
        wb.active[f"C{i}"].value = ls[i-2][2]
        wb.active[f"D{i}"].value = ls[i-2][3]
        wb.active[f"E{i}"].value = ls[i-2][4]
        wb.active[f"F{i}"].value = ls[i-2][5]
    wb.create_sheet(title='Специализация', index=1)
    wb.active = 1
    wb.active[f"A1"].value = "Номер сотрудника(id)"
    wb.active[f"B1"].value = "Вид работ(id)"
    ls = [[3,1],[10,1],[7,2],[4,2],[12,3],[3,4]]
    for i in range(2, 8):
        wb.active[f"A{i}"].value = ls[i-2][0]
        wb.active[f"B{i}"].value = ls[i-2][1]
    wb.create_sheet(title='Виды_работ', index=2)
    wb.active = 2
    wb.active[f"A1"].value = "id Работы"
    wb.active[f"B1"].value = "Название работы"
    wb.active[f"C1"].value = "Цена"
    wb.active[f"D1"].value = "Описание"
    ls = [[1, "Стрижка", 1200, "-"], [2,	"Маникюр", 1000, "-"],
        [3, "Мытьё", 300, "-"], [4, "Окраска", 2150, "-"]]
    for i in range(2, 6):
        wb.active[f"A{i}"].value = ls[i-2][0]
        wb.active[f"B{i}"].value = ls[i-2][1]
        wb.active[f"C{i}"].value = ls[i-2][2]
        wb.active[f"D{i}"].value = ls[i-2][3]
    wb.create_sheet(title='Клиенты', index=3)
    wb.active = 3
    wb.active[f"A1"].value = "Номер клиента"
    wb.active[f"B1"].value = "Имя"
    wb.active[f"C1"].value = "Фамилия"
    wb.active[f"D1"].value = "Телефон"
    wb.active[f"E1"].value = "Описание"
    ls = [[1,"Семён","Слепаков","+79324733228", "-"],
        [2,"Антон","Коротков","7583252356","-"],
        [3, "Александр","Пушкин","4283675892","-"],
        [4,"Максим", "Киселев","903469346","-"],
        [5,"Никита","Казаков","20620620","-"],
        [6,"Вася","Ловин","+7932534523","-"]]
    for i in range(2, 8):
        wb.active[f"A{i}"].value = ls[i-2][0]
        wb.active[f"B{i}"].value = ls[i-2][1]
        wb.active[f"C{i}"].value = ls[i-2][2]
        wb.active[f"D{i}"].value = ls[i-2][3]
        wb.active[f"E{i}"].value = ls[i-2][4]
    wb.create_sheet(title='бланк_заказа', index=4)
    wb.active = 4
    wb.active[f"A1"].value = "Номер заказа"
    wb.active[f"B1"].value = "Номер работы"
    wb.active[f"C1"].value = "Цена"
    wb.active[f"D1"].value = "Номер мастера"
    wb.active[f"E1"].value = "Номер клиента"
    wb.save('base.xlsx')
    logg("База данных, создана")
